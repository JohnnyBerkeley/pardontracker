"""
Export incidents.json and pardons.json to Excel.
Usage: python scripts/export_excel.py
Output: data/trump_money_map.xlsx
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent.parent
INCIDENTS_FILE = BASE / "data" / "incidents.json"
PARDONS_FILE = BASE / "data" / "pardons.json"
OUTPUT_FILE = BASE / "data" / "trump_money_map.xlsx"

RED = "C0392B"
DARK_GRAY = "2C3E50"
LIGHT_GRAY = "F2F3F4"
WHITE = "FFFFFF"
ORANGE = "E67E22"
BLUE = "2980B9"
GREEN = "27AE60"


def header_style(cell, bg=DARK_GRAY, fg=WHITE, bold=True):
    cell.font = Font(bold=bold, color=fg, name="Calibri", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def data_style(cell, wrap=True):
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    cell.font = Font(name="Calibri", size=9)


def stripe_row(ws, row, cols):
    if row % 2 == 0:
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def list_field(val):
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val) if val is not None else ""


def money(val):
    if val is None:
        return ""
    return f"${val:,.0f}"


# ── Load data ──────────────────────────────────────────────────────────────────

with open(INCIDENTS_FILE, encoding="utf-8") as f:
    idata = json.load(f)
with open(PARDONS_FILE, encoding="utf-8") as f:
    pdata = json.load(f)

incidents = idata["incidents"]
pardons = pdata["pardons"]

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Incidents
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Incidents"
ws1.freeze_panes = "A2"

cols1 = [
    "ID", "Date", "Date End", "Title", "Summary",
    "Categories", "Status", "Counterparty", "CP Type", "CP Country",
    "Beneficiary", "Actors", "Amount (USD)", "Amount Note",
    "Financial Connection", "Fin. Description",
    "Quid Pro Quo", "QPQ Gave", "QPQ Received",
    "Tags", "Last Updated"
]

for col, name in enumerate(cols1, 1):
    c = ws1.cell(row=1, column=col, value=name)
    header_style(c)

ws1.row_dimensions[1].height = 30

for row, inc in enumerate(incidents, 2):
    stripe_row(ws1, row, len(cols1))
    cp = inc.get("counterparty", {})
    fc = inc.get("financial_connection", {})
    qpq = inc.get("quid_pro_quo", {})

    values = [
        inc.get("id", ""),
        inc.get("date", ""),
        inc.get("date_end", ""),
        inc.get("title", ""),
        inc.get("summary", ""),
        list_field(inc.get("categories", [])),
        inc.get("status", ""),
        cp.get("name", ""),
        cp.get("type", ""),
        cp.get("country", ""),
        list_field(inc.get("beneficiary", [])),
        list_field(inc.get("actors", [])),
        inc.get("amount_usd"),
        inc.get("amount_note", ""),
        "Yes" if fc.get("exists") else ("No" if fc.get("exists") is False else "Unknown"),
        fc.get("description", ""),
        "Yes" if qpq.get("alleged") else "No",
        qpq.get("gave", ""),
        qpq.get("received", ""),
        list_field(inc.get("tags", [])),
        inc.get("last_updated", ""),
    ]
    for col, val in enumerate(values, 1):
        c = ws1.cell(row=row, column=col, value=val if not isinstance(val, (int, float)) else val)
        data_style(c)
        if col == 13 and val is not None:  # amount column — format as number
            c.number_format = '#,##0'

set_col_widths(ws1, [22, 10, 10, 35, 55, 30, 10, 25, 20, 8, 30, 35, 15, 35, 12, 45, 8, 45, 45, 30, 12])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Pardons
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pardons")
ws2.freeze_panes = "A2"

cols2 = [
    "ID", "Name", "Date", "Type", "Group", "Offense",
    "Original Sentence",
    "Fines Forgiven (USD)", "Restitution Forgiven (USD)", "Restitution Owed To",
    "Financial Connection", "Fin. Description",
    "Political Connection", "Pol. Description",
    "Notes"
]

for col, name in enumerate(cols2, 1):
    c = ws2.cell(row=1, column=col, value=name)
    header_style(c, bg=RED)

ws2.row_dimensions[1].height = 30

# Totals tracker
total_fines = 0
total_res_govt = 0
total_res_pvt = 0

for row, p in enumerate(pardons, 2):
    stripe_row(ws2, row, len(cols2))
    fc = p.get("financial_connection", {})
    pc = p.get("political_connection", {})

    fin_exists = fc.get("exists")
    pol_exists = pc.get("exists")

    fines = p.get("fines_forgiven_usd") or 0
    res = p.get("restitution_forgiven_usd") or 0
    owed_to = p.get("restitution_owed_to", "")
    total_fines += fines
    if owed_to == "private_victims":
        total_res_pvt += res
    else:
        total_res_govt += res

    values = [
        p.get("id", ""),
        p.get("name", ""),
        p.get("date", ""),
        p.get("type", ""),
        p.get("group", ""),
        p.get("offense", ""),
        p.get("original_sentence", ""),
        fines if fines else None,
        res if res else None,
        owed_to,
        "Yes" if fin_exists else ("No" if fin_exists is False else "Unknown"),
        fc.get("description", ""),
        "Yes" if pol_exists else ("No" if pol_exists is False else "Unknown"),
        pc.get("description", ""),
        p.get("notes", ""),
    ]
    for col, val in enumerate(values, 1):
        c = ws2.cell(row=row, column=col, value=val)
        data_style(c)
        if col in (8, 9) and val:
            c.number_format = '#,##0'
        if col == 11 and val == "Yes":
            c.font = Font(bold=True, color=RED, name="Calibri", size=9)
        if col == 13 and val == "Yes":
            c.font = Font(bold=True, color=ORANGE, name="Calibri", size=9)

# Totals row
totals_row = len(pardons) + 2
ws2.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True, name="Calibri")
ws2.cell(row=totals_row, column=7, value="Fines forgiven:").font = Font(bold=True, name="Calibri")
c_fin = ws2.cell(row=totals_row, column=8, value=total_fines)
c_fin.number_format = '#,##0'
c_fin.font = Font(bold=True, color=RED, name="Calibri")

ws2.cell(row=totals_row + 1, column=7, value="Restitution (govt):").font = Font(bold=True, name="Calibri")
c_rg = ws2.cell(row=totals_row + 1, column=9, value=total_res_govt)
c_rg.number_format = '#,##0'
c_rg.font = Font(bold=True, color=RED, name="Calibri")

ws2.cell(row=totals_row + 2, column=7, value="Restitution (private victims):").font = Font(bold=True, name="Calibri")
c_rp = ws2.cell(row=totals_row + 2, column=9, value=total_res_pvt)
c_rp.number_format = '#,##0'
c_rp.font = Font(bold=True, color=RED, name="Calibri")

ws2.cell(row=totals_row + 3, column=7, value="GRAND TOTAL FORGIVEN:").font = Font(bold=True, name="Calibri")
c_gt = ws2.cell(row=totals_row + 3, column=8, value=total_fines + total_res_govt + total_res_pvt)
c_gt.number_format = '#,##0'
c_gt.font = Font(bold=True, color=RED, name="Calibri", size=11)

set_col_widths(ws2, [28, 30, 10, 12, 18, 55, 40, 18, 18, 18, 12, 45, 12, 55, 55])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Pardons Sources
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Pardon Sources")
ws3.freeze_panes = "A2"

cols3 = ["Pardon ID", "Name", "Source Title", "URL", "Date", "Outlet", "Type"]
for col, name in enumerate(cols3, 1):
    c = ws3.cell(row=1, column=col, value=name)
    header_style(c, bg=RED)

row3 = 2
for p in pardons:
    for src in p.get("sources", []):
        stripe_row(ws3, row3, len(cols3))
        for col, val in enumerate([
            p.get("id", ""), p.get("name", ""),
            src.get("title", ""), src.get("url", ""),
            src.get("date", ""), src.get("outlet", ""), src.get("type", "")
        ], 1):
            c = ws3.cell(row=row3, column=col, value=val)
            data_style(c, wrap=False)
        row3 += 1

set_col_widths(ws3, [28, 30, 50, 70, 12, 25, 12])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Incident Sources
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Incident Sources")
ws4.freeze_panes = "A2"

cols4 = ["Incident ID", "Title", "Source Title", "URL", "Date", "Outlet", "Type"]
for col, name in enumerate(cols4, 1):
    c = ws4.cell(row=1, column=col, value=name)
    header_style(c)

row4 = 2
for inc in incidents:
    for src in inc.get("sources", []):
        stripe_row(ws4, row4, len(cols4))
        for col, val in enumerate([
            inc.get("id", ""), inc.get("title", ""),
            src.get("title", ""), src.get("url", ""),
            src.get("date", ""), src.get("outlet", ""), src.get("type", "")
        ], 1):
            c = ws4.cell(row=row4, column=col, value=val)
            data_style(c, wrap=False)
        row4 += 1

set_col_widths(ws4, [28, 40, 50, 70, 12, 25, 12])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Summary
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary")

def s(row, col, val, bold=False, color=None, size=10, bg=None):
    c = ws5.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, name="Calibri", size=size, color=color or "000000")
    c.alignment = Alignment(vertical="top")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

s(1, 1, "TRUMP MONEY MAP — SUMMARY", bold=True, size=14, color=RED)
s(2, 1, f"Generated: 2026-03-29 | Incidents: {len(incidents)} | Pardons: {len(pardons)}")

s(4, 1, "PARDONS — FORGIVEN TOTALS", bold=True, bg=DARK_GRAY, color=WHITE)
rows = [
    ("Fines & forfeitures forgiven (owed to government/taxpayers)", total_fines),
    ("Restitution forgiven — owed to government", total_res_govt),
    ("Restitution forgiven — owed to private victims", total_res_pvt),
    ("GRAND TOTAL TRACKED", total_fines + total_res_govt + total_res_pvt),
]
for i, (label, val) in enumerate(rows, 5):
    bold = (i == 8)
    lc = s(i, 1, label, bold=bold)
    vc = ws5.cell(row=i, column=2, value=val)
    vc.number_format = '#,##0'
    vc.font = Font(bold=bold, name="Calibri", size=10, color=RED if bold else "000000")
    vc.alignment = Alignment(vertical="top")

s(9, 1, "Note: CA Governor estimates ~$2B total; gap is from pardons with no disclosed amounts", color="666666")

from collections import defaultdict
s(11, 1, "INCIDENTS BY CATEGORY", bold=True, bg=DARK_GRAY, color=WHITE)
s(11, 2, "Count", bold=True, bg=DARK_GRAY, color=WHITE)
s(11, 3, "Total Amount (USD)", bold=True, bg=DARK_GRAY, color=WHITE)
s(11, 4, "Beneficiaries", bold=True, bg=DARK_GRAY, color=WHITE)

cat_totals = defaultdict(lambda: {"count": 0, "amount": 0, "bens": set()})
for inc in incidents:
    amt = inc.get("amount_usd") or 0
    bens = inc.get("beneficiary", [])
    for cat in inc.get("categories", []):
        cat_totals[cat]["count"] += 1
        cat_totals[cat]["amount"] += amt
        cat_totals[cat]["bens"].update(bens)

for i, (cat, v) in enumerate(sorted(cat_totals.items(), key=lambda x: -x[1]["amount"]), 12):
    s(i, 1, cat)
    ws5.cell(row=i, column=2, value=v["count"]).font = Font(name="Calibri", size=10)
    vc = ws5.cell(row=i, column=3, value=v["amount"] if v["amount"] else None)
    vc.number_format = '#,##0'
    vc.font = Font(name="Calibri", size=10)
    s(i, 4, ", ".join(sorted(v["bens"])))

s(30, 1, "NOTE: Category amounts overlap (one incident can appear in multiple categories). Do not sum.", color="666666")

ws5.column_dimensions["A"].width = 40
ws5.column_dimensions["B"].width = 10
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 55

wb.save(OUTPUT_FILE)
print(f"Saved: {OUTPUT_FILE}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"Pardons: {len(pardons)} | Incidents: {len(incidents)}")
print(f"Grand total forgiven: ${total_fines + total_res_govt + total_res_pvt:,.0f}")
